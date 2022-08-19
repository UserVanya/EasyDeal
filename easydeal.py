import PySimpleGUI as sg
import openpyxl
import os
import colorama
import subprocess
from functools import partial
import sys
import num2text
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime, date

common_price = 69832
privilaged_price = 15000
categories_exp = ["Потребности", "Транспорт", "Другие затраты", "Жилье", "Личные затраты", "Аптека"]
categories_inc = ["Жилье", "Остальное"]


def Confirm(hd):
    layout = [[sg.Text("Затраты на регистратора:"), sg.InputText(hd['Затраты на регистратора'])],
              [sg.Button("Сохранить"), sg.Button("Отмена"), sg.Button('Назад')]]
    event = None
    window = sg.Window('EasyDeal', layout)
    while event == None:
        event, values = window.read()
        if event in (None, 'Exit', 'Отмена'):
            window.close()
            sys.exit()
        if event in 'Назад':
            window.close()
            return None, event
        if event in 'Сохранить':
            hd['Затраты на регистратора'] = int(str(values[0]))
            hd['Договор подписан'] = "Да"
            window.close()
            return hd, event


def Add_successor(hd):
    new_hd = {}
    for key in hd:
        if key.startswith("Колво"):
            new_hd[key] = hd[key]
        elif key.startswith("Наследодатель"):
            new_hd[key] = hd["Фамилия"] + " " + hd["Имя"] + " " + hd["Отчество"]
        else:
            new_hd[key] = ""
    f = partial(field, new_hd)
    layout = [
        [f('Фамилия') + f('Имя') + f('Отчество')],
        [f('Дата рождения') + f('Место рождения')],
        [f('Серия') + f('Номер')],
        [f('Кем выдан')],
        [f('Дата выдачи') + f('Код подразделения')],
        [f('Область')],
        [sg.Text('Индекс регистрации'), sg.InputText(hd['Индекс регистрации'], key='-INDEXREG-', enable_events=True),
         sg.Text('Адрес регистрации'), sg.InputText(hd['Адрес регистрации'], key='-REG-', enable_events=True)],
        [sg.Text('Индекс проживания'), sg.InputText(hd['Индекс проживания'], key='-INDEXPROJ-', enable_events=True),
         sg.Text('Адрес проживания'), sg.InputText(hd['Адрес проживания'], key='-PROJ-', enable_events=True)],
        [sg.Text('АО: ' + str(hd['Колво АО'])), sg.Text('АП: ' + str(hd['Колво АП']))],
        [f('Наследодатель') + f('Наследник')],
        [f('Телефон')],
        [f('Комментарий')],
        [sg.Button("Сохранить"), sg.Button("Отмена")]
    ]
    window = sg.Window('EasyDeal', layout)
    event = None
    while event == None:  # The Event Loop
        event, values = window.read()
        if event in (None, 'Exit', 'Отмена'):
            window.close()
            sys.exit()
        if event == '-INDEXREG-':
            window['-INDEXPROJ-'].Update(values['-INDEXREG-'])
        if event == '-REG-':
            window['-PROJ-'].Update(values['-REG-'])
        if event == "Сохранить":
            v = []
            for el in values:
                v.append(values[el])
            first_part = ["", ""] + v[:15]
            second_part = v[15:]
            whole = first_part + [hd['Колво АО'], hd['Цена АО'], hd['Сумма АО'],
                                  hd['Колво АП'], hd['Цена АП'], hd['Сумма АП']]
            whole = whole + second_part
            whole = whole + [hd['Договор подписан'], hd['Затраты на регистратора']]
            window.close()
            return whole
        event = None


def Prepare(hd, descr, new_deal_num):
    layout = [
        [sg.Text(descr)],
        [sg.Output(size=(60, 30))]
    ]
    if hd["Номер договора"] == "" and hd["Дата договора"] == "":
        layout.append([sg.Text("Номер договора"), sg.InputText(str(new_deal_num), key='-NUMBER-'),
                       sg.Text("Дата договора"), sg.InputText(datetime.today().strftime("%d.%m.%Y"), key='-DATE-')])
    if hd['Колво АО'] != "0":
        layout.append([sg.Text("Колво АО на продажу(max " + hd['Колво АО'] + "):"),
                       sg.InputText("", key='-AO-', enable_events=True),
                       sg.Text("Цена 1 АО: "), sg.InputText("", key='-AOPRICE-', enable_events=True),
                       sg.Text("0", key='-AOSUM-', enable_events=True)])
    if hd['Колво АП'] != "0":
        layout.append([sg.Text("Колво АП на продажу(Макс " + hd['Колво АП'] + "):"),
                       sg.InputText("", key='-AP-', enable_events=True),
                       sg.Text("Цена 1 АП: "), sg.InputText("", key='-APPRICE-', enable_events=True),
                       sg.Text("0", key='-APSUM-', enable_events=True)])
    layout.append([sg.Button("Подготовить"), sg.Button("Отмена"), sg.Button("Назад")])
    window = sg.Window('EasyDeal', layout)
    event = None
    while event == None:  # The Event Loop
        event, values = window.read()
        if event in (None, 'Exit', 'Отмена'):
            window.close()
            sys.exit()
        if (event == '-AOPRICE-' or event == '-AO-') and str(values['-AO-']).isdigit() and str(
                values['-AOPRICE-']).isdigit():
            window['-AOSUM-'].Update(str(int(values['-AOPRICE-']) * int(values['-AO-'])))
        if (event == '-APPRICE-' or event == '-AP-') and str(values['-AP-']).isdigit() and str(
                values['-APPRICE-']).isdigit():
            window['-APSUM-'].Update(str(int(values['-APPRICE-']) * int(values['-AP-'])))
        if event == "Подготовить":
            if hd["Номер договора"] == "" and hd["Дата договора"] == "":
                hd["Номер договора"] = values['-NUMBER-']
                hd["Дата договора"] = values['-DATE-']
            if hd['Колво АО'] != "0":
                hd['Колво АО'] = str(int(values['-AO-']))
                hd['Цена АО'] = str(int(values['-AOPRICE-']))
                hd['Сумма АО'] = int(hd["Колво АО"]) * int(hd["Цена АО"])
            if hd['Колво АП'] != "0":
                hd['Колво АП'] = str(int(values['-AP-']))
                hd['Цена АП'] = str(int(values['-APPRICE-']))
                hd['Сумма АП'] = str(int(hd['Колво АП']) * int(hd['Цена АП']))
            window.close()
            return hd, event
        if event in ('Назад'):
            window.close()
            return None, event
        event = None


def Edit(hd):
    f = partial(field, hd)
    layout = [
        [f('Номер договора') + f('Дата договора')],
        [f('Фамилия') + f('Имя') + f('Отчество')],
        [f('Дата рождения') + f('Место рождения')],
        [f('Серия') + f('Номер')],
        [f('Кем выдан')],
        [f('Дата выдачи') + f('Код подразделения')],
        [f('Область')],
        [sg.Text('Индекс регистрации'), sg.InputText(hd['Индекс регистрации'], key='-INDEXREG-', enable_events=True),
         sg.Text('Адрес регистрации'), sg.InputText(hd['Адрес регистрации'], key='-REG-', enable_events=True)],
        [sg.Text('Индекс проживания'), sg.InputText(hd['Индекс проживания'], key='-INDEXPROJ-', enable_events=True),
         sg.Text('Адрес проживания'), sg.InputText(hd['Адрес проживания'], key='-PROJ-', enable_events=True)],
        [sg.Text('АО: ' + str(hd['Колво АО'])), sg.Text('АП: ' + str(hd['Колво АП']))],
        [f('Наследодатель') + f('Наследник')],
        [f('Телефон')],
        [f('Комментарий')],
        [sg.Button("Сохранить"), sg.Button("Отмена"), sg.Button("Назад")]
    ]
    window = sg.Window('EasyDeal', layout)
    event = None
    while event == None:  # The Event Loop
        event, values = window.read()
        if event in (None, 'Exit', 'Отмена'):
            window.close()
            sys.exit()
        if event == '-INDEXREG-':
            window['-INDEXPROJ-'].Update(values['-INDEXREG-'])
        if event == '-REG-':
            window['-PROJ-'].Update(values['-REG-'])
        if event in "Назад":
            window.close()
            return None, event
        if event == "Сохранить":
            v = []
            for el in values:
                v.append(values[el])
            first_part = v[:17]
            second_part = v[17:]
            whole = first_part + [hd['Колво АО'], hd['Цена АО'], hd['Сумма АО'],
                                  hd['Колво АП'], hd['Цена АП'], hd['Сумма АП']]
            whole = whole + second_part
            whole = whole + [hd['Договор подписан'], hd['Затраты на регистратора']]

            window.close()
            return whole, event
        event = None


def action_upon_holder():
    layout = [
        [sg.Button("Редактировать"), sg.Button("Подготовить документы"),
         sg.Button("Оформить документы"), sg.Button("Подтвердить подписание договора"), sg.Button("Назад")]
    ]
    window = sg.Window('EasyDeal', layout)
    event = None
    while event == None:  # The Event Loop
        event, values = window.read()
        if event in (None, 'Exit'):
            window.close()
            sys.exit()
        if event in (
        'Назад', 'Редактировать', 'Оформить документы', 'Подготовить документы', 'Подтвердить подписание договора'):
            window.close()
            return event


def field(holder, name):
    return [sg.Text(name + ":"), sg.InputText(holder[name])]


def Show_op(incomes, expenses):
    layout = [[sg.Button('Приходы'), sg.Button('Расходы')], [sg.Button('Назад')]]
    window = sg.Window("EasyDeal", layout)
    op = None
    while op == None:
        op, values = window.read()
        if op in (None, 'Exit'):
            window.close()
            sys.exit()
        if op in ('Приходы', 'Расходы'):
            window.close()
        if op in ('Назад'):
            window.close()
            return op
    if op == 'Приходы':
        headings = ["Дата добавления", "Категория", "Сумма"]
        layout = [[sg.Table(values=incomes, headings=headings,
                            auto_size_columns=True,
                            display_row_numbers=True,
                            justification='right',
                            num_rows=10)], [sg.Button("Назад")]]
    if op == 'Расходы':
        headings = ["Дата добавления", "Категория", "Название", "Сумма"]
        layout = [[sg.Table(values=expenses, headings=headings,
                            auto_size_columns=True,
                            display_row_numbers=True,
                            justification='right',
                            num_rows=10)], [sg.Button("Назад")]]
    window = sg.Window('EasyDeal', layout)
    event = None
    while event == None:
        event, values = window.read()
        if event in (None, 'Exit'):
            window.close()
            sys.exit()
        if event in ("Назад"):
            window.close()
            return event


def Add_op_menu():
    layout = [[sg.Button('Приход'), sg.Button('Расход')], [sg.Button('Назад')]]
    window = sg.Window("EasyDeal", layout)
    op = None
    while op == None:
        op, values = window.read()
        if op in (None, 'Exit'):
            window.close()
            sys.exit()
        if op in ('Приход', 'Расход'):
            window.close()
        if op in ('Назад'):
            window.close()
            return op, None
    if op == 'Расход':
        layout = [
            [sg.Text("Категория:"), sg.InputCombo(categories_exp), sg.Text("Название:"), sg.InputText(),
             sg.Text("Сумма:"), sg.InputText()],
            [sg.Button("Сохранить"), sg.Button("Отмена"), sg.Button('Назад')]
        ]
    if op == 'Приход':
        layout = [
            [sg.Text("Категория:"), sg.InputCombo(categories_inc),
             sg.Text("Сумма:"), sg.InputText()],
            [sg.Button("Сохранить"), sg.Button("Отмена"), sg.Button('Назад')]
        ]
    event = None
    window = sg.Window("EasyDeal", layout)
    while event == None:
        event, values = window.read()
        if event in (None, 'Exit', 'Отмена'):
            window.close()
            sys.exit()
        if event == "Сохранить":
            window.close()
            return op, values
        if event in "Назад":
            window.close()
            return event, None


def Cost_price(exp, data):
    bought_common = 0
    bought_privilaged = 0
    payouts_common = 0
    payouts_privilaged = 0
    exp_sum = 0
    reg = 0
    for holder in data:
        if holder["Договор подписан"] == "Да":
            bought_common = bought_common + int(holder["Колво АО"])
            bought_privilaged = bought_privilaged + int(holder["Колво АП"])
            payouts_common = payouts_common + int(holder["Сумма АО"])
            payouts_privilaged = payouts_privilaged + int(holder["Сумма АП"])
            reg = reg + int(holder["Затраты на регистратора"])
    payouts_with_reg = payouts_common + payouts_privilaged + reg
    for el in exp:
        if el[2] != 'Личные затраты':
            exp_sum = exp_sum + int(el[3])
    part = (exp_sum + reg) / (payouts_common + payouts_privilaged)
    cost_common = (part * payouts_common + payouts_common) / bought_common
    cost_privilaged = (part * payouts_privilaged + payouts_privilaged) / bought_privilaged
    return (exp_sum + payouts_with_reg) / (
                bought_common + bought_privilaged), bought_common, bought_privilaged, cost_common, cost_privilaged


def Menu(data, cost_price, bought_common, bought_privilaged, cost_common, cost_privilaged):
    layout = [
        [sg.Button("Добавить приход/расход"), sg.Button("Показать приходы/расходы")],
        [sg.Button("Поиск акционера"), sg.Button("Сформировать отчет"), sg.Button("Выйти")],
        [sg.Text("Куплено бумаг АО:" + str(bought_common) + "   АП:" + str(bought_privilaged))],
        [sg.Text("Себестоимость бумаги: " + str(cost_price))],
        [sg.Text("Себестоимость 1 превилегированной акции: " + str(cost_privilaged))],
        [sg.Text("Себестоимость 1 обыкновенной акции: " + str(cost_common))]
    ]
    event = None
    window = sg.Window('EasyDeal', layout)
    while event == None:
        event, values = window.read()
        if event in (None, 'Exit', 'Выйти'):
            window.close()
            sys.exit()
        if event in ("Добавить приход/расход", "Показать приходы/расходы", "Поиск акционера", "Сформировать отчет"):
            window.close()
            return event


def Find_holder(descr):
    layout = [
        [sg.Text('ФИО'), sg.InputText('', key='-INPUT-', enable_events=True),
         sg.Checkbox('Создать акционера на \n основании данного наследодателя', key='-CB-'),
         sg.Button("Назад"), sg.Button('Отмена')],
        [sg.Listbox(descr, key='-INPUTCOMBO-', size=(80, 20), enable_events=True), sg.Button("Выбрать")]
    ]
    window = sg.Window("EasyDeal", layout)
    event = None
    while event == None:
        event, values = window.read()
        if event in (None, 'Exit', 'Отмена'):
            window.close()
            sys.exit()
        if event in ('Назад'):
            window.close()
            return None, None, event
        if event == "Выбрать":
            window.close()
            return values['-INPUT-'], values['-CB-'], event
        if event == '-INPUT-':
            text = values['-INPUT-']
            fetches = []
            if text != "" and (len(text) > 1 or text[len(text) - 1] == ' '):
                for el in descr:
                    if str.lower(el).startswith(str.lower(text)):
                        fetches.append(el)
                window['-INPUTCOMBO-'].Update(values=fetches)
            else:
                window['-INPUTCOMBO-'].Update(values=descr)
            event = None
        if event == '-INPUTCOMBO-':
            window['-INPUT-'].Update(values['-INPUTCOMBO-'][0])
            event = None


def Get_sheets():
    layout = [
        [sg.Text('База данных акционеров(.xlsx)'),
         sg.InputText(''), sg.FileBrowse()],
        [sg.Text('Название листа(БД):'), sg.InputText('Лист1'), sg.Text('Название листа(Бух.учет):'),
         sg.InputText('Лист2')],
        [sg.Output(size=(88, 20))],
        [sg.Button("Далее"), sg.Button("Отменить")]
    ]
    window = sg.Window('EasyDeal', layout)
    event = None
    wb, sheet1, sheet2 = None, None, None
    exception_caught = False
    values = None
    while event == None:  # The Event Loop
        event, values = window.read()
        if event in (None, 'Exit', 'Отменить'):
            window.close()
            sys.exit()
        if event == 'Далее':
            try:
                exception_caught = True
                wb = openpyxl.load_workbook(filename=values[0])
                sheet1 = wb[values[1]]
                sheet2 = wb[values[2]]
                exception_caught = False
            except InvalidFileException:
                sg.popup("Неверный формат файла")
            except FileNotFoundError:
                sg.popup("Неверный путь или имя файла")
            except KeyError:
                sg.popup("Неверное название листа")
            if exception_caught == True:
                event = None
                exception_caught = False
    window.Close()
    return sheet1, sheet2, wb, values[0]


colorama.init()
auxiliary = ["Номер договора",  # 1
             "Дата договора",  # 2
             "Фамилия",  # 3
             "Имя",  # 4
             "Отчество",  # 5
             "Дата рождения",  # 6
             "Место рождения",  # 7
             "Серия",  # 8
             "Номер",  # 9
             "Кем выдан",  # 10
             "Дата выдачи",  # 11
             "Код подразделения",  # 12
             "Область",  # 13
             "Индекс регистрации",  # 14
             "Адрес регистрации",  # 15
             "Индекс проживания",  # 16
             "Адрес проживания",  # 17
             "Колво АО",  # 18
             "Цена АО",  # 19
             "Сумма АО",  # 20
             "Колво АП",  # 21
             "Цена АП",  # 22
             "Сумма АП",  # 23
             "Наследодатель",  # 24
             "Наследник",  # 25
             "Телефон",  # 26
             "Комментарий",  # 27
             "Договор подписан",  # 28
             "Затраты на регистратора"  # 29
             ]
aux = {}
i = 1
for key in auxiliary:
    val = key.title().replace(' ', "")
    aux[key] = (val, i)
    i = i + 1
sheet, sheet2, wb, wb_full_name = Get_sheets()
none_row_n = 1
i = 1
while sheet.cell(row=i, column=aux["Фамилия"][1]).value != None:
    none_row_n = none_row_n + 1
    i = i + 1
data = []
descr = []
for i in range(1, none_row_n):
    data.append({})
    for key in aux:
        if key == ("Колво АО" or "Колво ПА") and (
                sheet.cell(row=i, column=aux[key][1]).value == None or sheet.cell(row=i, column=aux[key][
            1]).value == "None") or sheet.cell(row=i, column=aux[key][1]).value == "":
            data[i - 1][key] = "0"
        elif key == ("Сумма АО" or "Сумма ПА") and (
                sheet.cell(row=i, column=aux[key][1]).value == None or sheet.cell(row=i, column=aux[key][
            1]).value == "None") or sheet.cell(row=i, column=aux[key][1]).value == "":
            data[i - 1][key] = "0"
        elif key.startswith("Дата"):
            if sheet.cell(row=i, column=aux[key][1]).value != None:
                data[i - 1][key] = sheet.cell(row=i, column=aux[key][1]).value.strftime("%d.%m.%Y")
            else:
                data[i - 1][key] = ""
        else:
            if sheet.cell(row=i, column=aux[key][1]).value == None:
                data[i - 1][key] = ""
            else:
                data[i - 1][key] = str(sheet.cell(row=i, column=aux[key][1]).value)
    descr.append(data[i - 1]["Фамилия"] + " " + data[i - 1]["Имя"] + " " + data[i - 1]["Отчество"] + " "
                 + data[i - 1]["Дата рождения"] + " " + data[i - 1]["Серия"] + " "
                 + data[i - 1]["Номер"] + " "
                 + " AО:" + data[i - 1]["Колво АО"] + " АП:" + data[i - 1]["Колво АП"])
    if data[i - 1]["Договор подписан"] == 'Да':
        descr[i - 1] = descr[i - 1] + "(Подписан)"
incomes = []
expenses = []
none_row_n1 = 2
while sheet2.cell(row=none_row_n1, column=1).value != None:
    incomes.append([])
    for i in range(1, 4):
        incomes[none_row_n1 - 2].append(sheet2.cell(row=none_row_n1, column=i).value)
    none_row_n1 = none_row_n1 + 1

none_row_n2 = 2
while sheet2.cell(row=none_row_n2, column=4).value != None:
    expenses.append([])
    for i in range(4, 8):
        expenses[none_row_n2 - 2].append(sheet2.cell(row=none_row_n2, column=i).value)
    none_row_n2 = none_row_n2 + 1
after_menu_event = None
while True:
    cost_price, bought_common, bought_privilaged, cost_common, cost_privilaged = Cost_price(expenses, data)
    menu_action = Menu(data, cost_price, bought_common, bought_privilaged, cost_common, cost_privilaged)
    if menu_action == "Добавить приход/расход":
        op, values = Add_op_menu()
        date = datetime.now()
        if op == 'Назад':
            continue
        if op == 'Приход':
            incomes.append([])
            incomes[len(incomes) - 1].append(date)
            incomes[len(incomes) - 1].append(values[0])
            incomes[len(incomes) - 1].append(values[1])
            sheet2.cell(row=none_row_n1, column=1).value = date
            sheet2.cell(row=none_row_n1, column=2).value = values[0]
            sheet2.cell(row=none_row_n1, column=3).value = int(values[1])
            none_row_n1 = none_row_n1 + 1
        if op == 'Расход':
            expenses.append([])
            expenses[len(expenses) - 1].append(date)
            expenses[len(expenses) - 1].append(values[0])
            expenses[len(expenses) - 1].append(values[1])
            expenses[len(expenses) - 1].append(values[2])
            sheet2.cell(row=none_row_n2, column=4).value = date
            sheet2.cell(row=none_row_n2, column=5).value = values[0]
            sheet2.cell(row=none_row_n2, column=6).value = values[1]
            sheet2.cell(row=none_row_n2, column=7).value = int(values[2])
            none_row_n2 = none_row_n2 + 1
        is_saved = False
        while is_saved != True:
            try:
                wb.save(wb_full_name)
                is_saved = True
            except:
                sg.popup("Закройте пожалуйста файл")
                is_saved = False
    elif menu_action == 'Показать приходы/расходы':
        after_menu_event = Show_op(incomes, expenses)
        if after_menu_event == "Назад":
            continue
    elif menu_action == 'Сформировать отчет':
        report_wb = openpyxl.Workbook()
        report_ws = report_wb.active
        inc_sum = {}
        exp_sum = {}
        inc_sum['Приходы'] = 0
        exp_sum['Расходы'] = 0
        for el in categories_exp:
            exp_sum[el] = 0
        for el in categories_inc:
            inc_sum[el] = 0
        for el in incomes:
            inc_sum[str(el[1])] = inc_sum[el[1]] + el[2]
            inc_sum['Приходы'] = inc_sum['Приходы'] + el[2]
        for el in expenses:
            exp_sum[str(el[1])] = exp_sum[el[1]] + int(el[3])
            if el[1] != 'Личные затраты':
                exp_sum['Расходы'] = exp_sum['Расходы'] + int(el[3])
        i = 1
        for key in inc_sum:
            report_ws.cell(row=i, column=1).value = key
            report_ws.cell(row=i, column=2).value = inc_sum[key]
            i = i + 1
        i = 1
        for key in exp_sum:
            if key != "Личные затраты":
                report_ws.cell(row=i, column=4).value = key
                report_ws.cell(row=i, column=5).value = exp_sum[key]
                i = i + 1
        report_ws.cell(row=len(exp_sum) + 2, column=4).value = "Личные затраты(Зеленин)"
        report_ws.cell(row=len(exp_sum) + 2, column=5).value = exp_sum['Личные затраты']
        report_ws.cell(row=len(exp_sum) + 4, column=4).value = "Куплено АО"
        report_ws.cell(row=len(exp_sum) + 4, column=5).value = bought_common
        report_ws.cell(row=len(exp_sum) + 5, column=4).value = "Куплено АП"
        report_ws.cell(row=len(exp_sum) + 5, column=5).value = bought_privilaged
        payouts_common = 0
        payouts_privilaged = 0
        reg = 0
        for my_h in data:
            if my_h["Договор подписан"] == "Да":
                payouts_common = payouts_common + int(my_h["Сумма АО"])
                payouts_privilaged = payouts_privilaged + int(my_h["Сумма АП"])
                reg = reg + int(my_h['Затраты на регистратора'])
        report_ws.cell(row=len(exp_sum) + 7, column=4).value = "Сумма АО"
        report_ws.cell(row=len(exp_sum) + 7, column=5).value = payouts_common
        report_ws.cell(row=len(exp_sum) + 8, column=4).value = "Сумма АП"
        report_ws.cell(row=len(exp_sum) + 8, column=5).value = payouts_privilaged
        report_ws.cell(row=len(exp_sum) + 9, column=4).value = "Затраты на регистратора"
        report_ws.cell(row=len(exp_sum) + 9, column=5).value = reg

        report_ws.cell(row=1, column=7).value = "Себестоимость общая"
        report_ws.cell(row=1, column=8).value = cost_price
        report_ws.cell(row=2, column=7).value = "Себестоимость АО"
        report_ws.cell(row=2, column=8).value = cost_common
        report_ws.cell(row=3, column=7).value = "Себестоимость АП"
        report_ws.cell(row=3, column=8).value = cost_privilaged

        report_wb.save("Отчеты\Отчет " + datetime.today().strftime("%d.%m.%Y") + ".xlsx")
        os.startfile("Отчеты")
    elif menu_action == 'Поиск акционера':
        while True:
            holder, is_successor, after_menu_event = Find_holder(descr)
            if after_menu_event == "Назад":
                break
            ind = descr.index(holder)
            hd = data[ind]
            if not (hd['Колво АО'] == "0" and hd['Колво АП']) == "0":
                if is_successor:
                    new_holder = Add_successor(hd)
                    data[ind]["Колво АП"] = 0
                    data[ind]["Колво АО"] = 0
                    sheet.cell(row=ind + 1, column=25).value = new_holder[2] + " " + new_holder[3] + " " + new_holder[4]
                    for k in range(18, 24):
                        sheet.cell(row=ind + 1, column=k).value = 0
                    data[ind]["Наследник"] = sheet.cell(row=ind + 1, column=25).value
                    i = 0
                    j = len(data)
                    data.append({})
                    for el in aux:
                        data[j][el] = new_holder[i]
                        i = i + 1
                        if el.startswith("Дата") and data[j][el] != "":
                            sheet.cell(row=j + 1, column=i).value = datetime.strptime(data[j][el], "%d.%m.%Y")
                        elif el.find("АО") != -1 or el.find("АП") != -1:
                            sheet.cell(row=j + 1, column=i).value = int(data[j][el])
                        else:
                            sheet.cell(row=j + 1, column=i).value = data[j][el]

                    hd = data[j]
                    descr.append(hd["Фамилия"] + " " + hd["Имя"] + " " + hd["Отчество"] + " "
                                 + hd["Дата рождения"] + " " + hd["Серия"] + " "
                                 + hd["Номер"] + " " + " AО:" + hd["Колво АО"] + " АП:" + hd["Колво АП"])
                    ind = j
                    is_saved = False
                    while is_saved != True:
                        try:
                            wb.save(wb_full_name)
                            is_saved = True
                        except:
                            sg.popup("Закройте пожалуйста файл")
                            is_saved = False
                while True:
                    action = action_upon_holder()
                    if action == "Назад":
                        break
                    if action == "Редактировать":
                        edited, event = Edit(hd)
                        if event == "Назад":
                            continue
                        i = 0
                        for el in aux:
                            data[ind][el] = edited[i]
                            i = i + 1
                            if el.startswith("Дата") and data[ind][el] != "":
                                sheet.cell(row=ind + 1, column=i).value = datetime.strptime(data[ind][el], "%d.%m.%Y")
                            elif el.find("АО") != -1 or el.find("АП") != -1:
                                sheet.cell(row=ind + 1, column=i).value = int(data[ind][el])
                            else:
                                sheet.cell(row=ind + 1, column=i).value = data[ind][el]
                        is_saved = False
                        while is_saved != True:
                            try:
                                wb.save(wb_full_name)
                                is_saved = True
                            except:
                                sg.popup("Закройте пожалуйста файл")
                                is_saved = False
                    elif action == "Подготовить документы":
                        deals_num = []
                        for el in data:
                            if el["Номер договора"] != "":
                                deals_num.append(int(el["Номер договора"]))
                        new_deal_num = 1
                        if len(deals_num) != 0:
                            new_deal_num = max(max(deals_num), 0) + 1
                        prepared, event = Prepare(hd, holder, new_deal_num)
                        if event == 'Назад':
                            continue
                        sheet.cell(row=ind + 1, column=18).value = int(sheet.cell(row=ind + 1, column=18).value) - int(
                            prepared["Колво АО"])
                        sheet.cell(row=ind + 1, column=21).value = int(sheet.cell(row=ind + 1, column=21).value) - int(
                            prepared["Колво АП"])
                        if sheet.cell(row=ind + 1, column=18).value == 0 and sheet.cell(row=ind + 1,
                                                                                        column=21).value == 0:
                            sheet.cell(row=ind + 1, column=1).value = ""
                            sheet.cell(row=ind + 1, column=2).value = ""
                        i = 0
                        j = len(data)
                        data.append({})
                        for el in aux:
                            if el.startswith("Дата") and prepared[el] != "":
                                data[j][el] = datetime.strptime(prepared[el], '%d.%m.%Y').strftime("%d.%m.%Y")
                            else:
                                data[j][el] = prepared[el]
                            i = i + 1
                            if el.find("АО") != -1 or el.find("АП") != -1:
                                sheet.cell(row=j + 1, column=i).value = int(data[j][el])
                            elif el.startswith("Дата") and data[ind][el] != "":
                                sheet.cell(row=j + 1, column=i).value = datetime.strptime(data[j][el], "%d.%m.%Y")
                            else:
                                sheet.cell(row=j + 1, column=i).value = data[j][el]
                        data[ind]['Колво АО'] = sheet.cell(row=ind + 1, column=18).value
                        data[ind]['Колво АП'] = sheet.cell(row=ind + 1, column=21).value
                        hd = data[j]
                        descr.append(hd["Фамилия"] + " " + hd["Имя"] + " " + hd["Отчество"] + " "
                                     + hd["Дата рождения"] + " " + hd["Серия"] + " "
                                     + hd["Номер"] + " " + " AО:" + hd["Колво АО"] + " АП:" + hd["Колво АП"])
                        ind = j
                        is_saved = False
                        while is_saved != True:
                            try:
                                wb.save(wb_full_name)
                                is_saved = True
                            except:
                                sg.popup("Закройте пожалуйста файл")
                                is_saved = False
                    elif action == "Оформить документы":
                        context = {}
                        for key in aux:
                            if key.startswith("Дата"):
                                context[aux[key][0]] = hd[key]
                            else:
                                context[aux[key][0]] = hd[key]

                        if hd['Колво АО'] != "0":
                            context['ПрописьюАо'] = num2text.num2text(int(hd['Колво АО']),
                                                                      main_units=((u'', u'', u''), 'f'))
                            context['ПрописьюСуммаАо'] = num2text.num2text(int(hd['Сумма АО']),
                                                                           main_units=((u'', u'', u''), 'm'))
                        if hd['Колво АП'] != "0":
                            context['ПрописьюАп'] = num2text.num2text(int(hd['Колво АП']),
                                                                      main_units=((u'', u'', u''), 'f'))
                            context['ПрописьюСуммаАп'] = num2text.num2text(int(hd['Сумма АП']),
                                                                           main_units=((u'', u'', u''), 'm'))

                        from docxtpl import DocxTemplate

                        fio = hd["Фамилия"] + " " + hd["Имя"] + " " + hd["Отчество"] + " АО " + hd[
                            "Колво АО"] + " АП " + hd["Колво АП"]
                        folder_name = "Акционеры/" + fio
                        if os.path.exists(folder_name):
                            os.remove(folder_name)
                        os.mkdir(folder_name)
                        doc = DocxTemplate("Шаблоны/АНКЕТА.docx")
                        doc.render(context)
                        doc.save(folder_name + "/АНКЕТА " + fio + ".docx")
                        if hd["Колво АП"] != "0":
                            doc = DocxTemplate("Шаблоны/ДКП ПРЕВ.docx")
                            doc.render(context)
                            doc.save(folder_name + "/ДКП ПРЕВ" + fio + ".docx")
                            doc = DocxTemplate("Шаблоны/РАСП ПРЕВ.docx")
                            doc.render(context)
                            doc.save(folder_name + "/РАСП ПРЕВ " + fio + ".docx")
                        if hd["Колво АО"] != "0":
                            doc = DocxTemplate("Шаблоны/ДКП ОБЫЧКА.docx")
                            doc.render(context)
                            doc.save(folder_name + "/ДКП ОБЫЧКА" + fio + ".docx")
                            doc = DocxTemplate("Шаблоны/РАСП ОБЫЧКА.docx")
                            doc.render(context)
                            doc.save(folder_name + "/РАСП ОБЫЧКА " + fio + ".docx")
                        os.startfile("Акционеры")
                    else:  # Подтвердить подписание договора
                        hd, event = Confirm(hd)
                        if event == 'Назад':
                            continue
                        sheet.cell(row=ind + 1, column=28).value = hd["Договор подписан"]
                        sheet.cell(row=ind + 1, column=29).value = int(hd["Затраты на регистратора"])
                        data[ind]["Договор подписан"] = hd["Договор подписан"]
                        data[ind]["Затраты на регистратора"] = hd["Затраты на регистратора"]
                        is_saved = False
                        while is_saved != True:
                            try:
                                wb.save(wb_full_name)
                                is_saved = True
                            except:
                                sg.popup("Закройте пожалуйста файл")
                                is_saved = False
            else:
                sg.popup("У данного акционера больше нет акций")
                break